#!/usr/bin/env python3
"""
storage/scripts/generate_margenes.py

Genera el reporte Excel de Márgenes y Bono Mensual.
Recibe JSON por stdin, escribe el .xlsx en output_path.

Dependencias:
    pip install openpyxl --break-system-packages

Hojas del workbook:
    1. Resumen Bono  → KPIs del período + semáforo global
    2. Márgenes      → Tabla detallada por artículo con colores de semáforo
"""

import sys
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers as xl_numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ── Paleta de colores ─────────────────────────────────────────────────────────
CLR_PRIMARY     = "1A56DB"  # Azul primario (brand)
CLR_HEADER_BG   = "0F172A"  # Sidebar dark
CLR_HEADER_FG   = "FFFFFF"
CLR_SUBHEADER   = "EFF6FF"
CLR_VERDE_BG    = "DCFCE7"
CLR_VERDE_FG    = "15803D"
CLR_AMARILLO_BG = "FEF3C7"
CLR_AMARILLO_FG = "92400E"
CLR_ROJO_BG     = "FEE2E2"
CLR_ROJO_FG     = "B91C1C"
CLR_NEGATIVO_BG = "EDE9FE"
CLR_NEGATIVO_FG = "5B21B6"
CLR_TOTAL_BG    = "F1F5F9"
CLR_BORDER      = "E2E8F0"

# ── Estilos reutilizables ────────────────────────────────────────────────────
def header_font(size=11, bold=True, color=CLR_HEADER_FG):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="0F172A"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    side = Side(style="thin", color=CLR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def border_bottom():
    return Border(bottom=Side(style="thin", color=CLR_BORDER))

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def vcenter():
    return Alignment(horizontal="left", vertical="center")

# ── Formato numérico ─────────────────────────────────────────────────────────
FMT_MONEY   = '#,##0.00'
FMT_MONEY_0 = '#,##0'
FMT_PCT     = '0.00%'
FMT_PCT_1   = '0.0%'

# ─────────────────────────────────────────────────────────────────────────────
# HOJA 1 — Resumen Bono Mensual
# ─────────────────────────────────────────────────────────────────────────────

def build_hoja_resumen(ws, resumen: dict, client_name: str, from_: str, to_: str,
                        cost_field: str, currency: str, iva_rate: float):

    ws.title = "Resumen Bono"
    ws.sheet_view.showGridLines = False

    # ── Título principal ──────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{client_name} — Reporte de Bono Mensual"
    ws["A1"].font      = Font(name="Arial", size=14, bold=True, color=CLR_HEADER_FG)
    ws["A1"].fill      = fill(CLR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Subtítulo período ─────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Período: {from_}  →  {to_}  |  Campo de costo: {cost_field}  |  IVA: {iva_rate}%"
    ws["A2"].font      = Font(name="Arial", size=9, color="64748B")
    ws["A2"].fill      = fill("F8FAFC")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8  # Espacio

    # ── Función auxiliar para filas KPI ──────────────────────────────────
    def kpi_row(row, label, value, fmt=FMT_MONEY, bg="FFFFFF", bold=False, fg_val="0F172A"):
        ws.cell(row=row, column=1).value     = label
        ws.cell(row=row, column=1).font      = Font(name="Arial", size=10, bold=bold, color="64748B")
        ws.cell(row=row, column=1).alignment = vcenter()
        ws.cell(row=row, column=1).fill      = fill(bg)

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=2)
        cell.value           = value
        cell.number_format   = fmt
        cell.font            = Font(name="Arial", size=11, bold=True, color=fg_val)
        cell.alignment       = Alignment(horizontal="right", vertical="center")
        cell.fill            = fill(bg)
        ws.row_dimensions[row].height = 22

    # ── KPIs financieros ──────────────────────────────────────────────────
    r = 4
    kpi_row(r,   "Total Facturado del Período",
            resumen.get("total_facturado", 0), bold=True)
    r += 1
    if resumen.get("iva_excluido"):
        kpi_row(r, f"  (−) IVA {iva_rate}% excluido",
                -resumen.get("iva_monto", 0), fg_val="7C3AED", bg="EDE9FE")
        r += 1
        kpi_row(r, "Base de cálculo (sin IVA)",
                resumen.get("total_base", 0), bold=True, bg="EFF6FF", fg_val=CLR_PRIMARY)
        r += 1

    kpi_row(r,   "(−) Costo Total de lo Vendido",
            -resumen.get("costo_total", 0), fg_val="B91C1C", bg="FEE2E2")
    r += 1

    # Línea separadora
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = Border(top=Side(style="medium", color=CLR_PRIMARY))
    r += 1

    # Ganancia neta — destacada
    gn = resumen.get("ganancia_neta", 0)
    gn_bg = CLR_VERDE_BG if gn >= 0 else CLR_ROJO_BG
    gn_fg = CLR_VERDE_FG if gn >= 0 else CLR_ROJO_FG
    kpi_row(r, "GANANCIA NETA DEL PERÍODO", gn,
            bold=True, bg=gn_bg, fg_val=gn_fg)
    ws.row_dimensions[r].height = 28
    ws.cell(row=r, column=1).font = Font(name="Arial", size=11, bold=True, color=gn_fg)
    r += 1

    margen_pct = resumen.get("margen_neto_pct", 0) / 100
    kpi_row(r, "Margen Neto del Período (%)", margen_pct, fmt=FMT_PCT,
            bold=True, bg=gn_bg, fg_val=gn_fg)
    ws.row_dimensions[r].height = 26
    r += 2

    # ── Conteo de artículos por semáforo ─────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1).value     = "Distribución de Artículos por Margen"
    ws.cell(row=r, column=1).font      = header_font(size=10, color=CLR_HEADER_BG)
    ws.cell(row=r, column=1).fill      = fill("EFF6FF")
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    semaforos = resumen.get("semaforo_conteo", {})
    total_art  = resumen.get("total_articulos", 1) or 1

    sem_data = [
        ("Margen Alto  (≥ umbral verde)",   semaforos.get("verde",    0), CLR_VERDE_BG,    CLR_VERDE_FG),
        ("Margen Medio (entre umbrales)",   semaforos.get("amarillo", 0), CLR_AMARILLO_BG, CLR_AMARILLO_FG),
        ("Margen Bajo  (< umbral rojo)",    semaforos.get("rojo",     0), CLR_ROJO_BG,     CLR_ROJO_FG),
        ("Margen Negativo",                 semaforos.get("negativos",0), CLR_NEGATIVO_BG, CLR_NEGATIVO_FG),
    ]

    for label, count, bg, fg in sem_data:
        ws.cell(row=r, column=1).value     = label
        ws.cell(row=r, column=1).font      = body_font(color=fg, bold=True)
        ws.cell(row=r, column=1).fill      = fill(bg)
        ws.cell(row=r, column=1).alignment = vcenter()
        ws.cell(row=r, column=2).value     = count
        ws.cell(row=r, column=2).font      = body_font(bold=True, color=fg)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).fill      = fill(bg)
        ws.cell(row=r, column=3).value     = count / total_art
        ws.cell(row=r, column=3).number_format = FMT_PCT_1
        ws.cell(row=r, column=3).font      = body_font(color=fg)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r, column=3).fill      = fill(bg)
        ws.row_dimensions[r].height = 20
        r += 1

    # ── Nota al pie ───────────────────────────────────────────────────────
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1).value = (
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
        f"Campo de costo: {cost_field} | "
        f"IVA {'excluido' if resumen.get('iva_excluido') else 'incluido'} en base de cálculo"
    )
    ws.cell(row=r, column=1).font      = Font(name="Arial", size=8, color="94A3B8", italic=True)
    ws.cell(row=r, column=1).alignment = vcenter()

    # ── Anchos de columna ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 10

    # ── Freeze panes ─────────────────────────────────────────────────────
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 2 — Detalle de Márgenes
# ─────────────────────────────────────────────────────────────────────────────

COLS = [
    ("Código",            12, "codigo"),
    ("Descripción",       42, "descripcion"),
    ("Precio Venta",      14, "precio_calculo"),
    ("Costo",             14, "costo"),
    ("Margen $",          14, "margen_monto"),
    ("Margen %",          10, "margen_pct"),
    ("Uds. Vendidas",     13, "unidades_vendidas"),
    ("Ingreso Total",     16, None),   # fórmula: precio * uds
    ("Costo Total",       14, None),   # fórmula: costo * uds
    ("Ganancia",          14, None),   # fórmula: ingreso - costo
    ("Alerta",            10, "semaforo"),
]

def build_hoja_margenes(ws, margenes: list, currency: str,
                         alert_red: float, alert_yellow: float,
                         excluir_iva: bool, iva_rate: float):

    ws.title = "Márgenes Detalle"
    ws.sheet_view.showGridLines = False

    n_cols = len(COLS)

    # ── Cabecera ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1).value     = f"Detalle de Márgenes por Artículo{' (precios sin IVA)' if excluir_iva else ''}"
    ws.cell(1, 1).font      = header_font(size=12)
    ws.cell(1, 1).fill      = fill(CLR_HEADER_BG)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Nota IVA ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    iva_nota = f"IVA ({iva_rate}%) EXCLUIDO del precio de venta para cálculo de margen" if excluir_iva \
               else f"Precio de venta incluye IVA ({iva_rate}%) — margen calculado sobre precio bruto"
    ws.cell(2, 1).value     = iva_nota
    ws.cell(2, 1).font      = Font(name="Arial", size=9,
                                   color=CLR_NEGATIVO_FG if excluir_iva else "64748B",
                                   italic=True)
    ws.cell(2, 1).fill      = fill("EDE9FE" if excluir_iva else "F8FAFC")
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Leyenda de colores ────────────────────────────────────────────────
    ws.row_dimensions[3].height = 14

    # ── Encabezados de columna ────────────────────────────────────────────
    for col_idx, (col_label, col_w, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value           = col_label
        cell.font            = header_font(size=9)
        cell.fill            = fill("1E293B")
        cell.alignment       = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border          = border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    ws.row_dimensions[4].height = 22

    # Freeze hasta cabecera
    ws.freeze_panes = "A5"

    # ── Filas de datos ────────────────────────────────────────────────────
    sem_styles = {
        "verde":    (CLR_VERDE_BG,    CLR_VERDE_FG,    "▲ Alto"),
        "amarillo": (CLR_AMARILLO_BG, CLR_AMARILLO_FG, "◆ Medio"),
        "rojo":     (CLR_ROJO_BG,     CLR_ROJO_FG,     "▼ Bajo"),
    }

    for row_offset, item in enumerate(margenes, start=5):
        r     = row_offset
        sem   = item.get("semaforo", "rojo")
        negat = item.get("es_negativo", False)

        if negat:
            row_bg = CLR_NEGATIVO_BG
        else:
            row_bg = sem_styles.get(sem, (CLR_ROJO_BG,))[0]

        def data_cell(col, value, fmt=None, bold=False, align="left", fg="0F172A"):
            c = ws.cell(row=r, column=col)
            c.value     = value
            c.font      = Font(name="Arial", size=9, bold=bold,
                               color=(sem_styles[sem][1] if negat else fg))
            c.fill      = fill(row_bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border    = border_bottom()
            if fmt:
                c.number_format = fmt

        # Columnas fijas
        data_cell(1, item.get("codigo", ""),         align="left")
        data_cell(2, item.get("descripcion", ""),     align="left")
        data_cell(3, item.get("precio_calculo", 0),   fmt=FMT_MONEY, align="right")
        data_cell(4, item.get("costo", 0),            fmt=FMT_MONEY, align="right")

        # Margen $ — colorear si negativo
        monto_fg = CLR_ROJO_FG if negat else CLR_VERDE_FG
        c5 = ws.cell(row=r, column=5)
        c5.value          = item.get("margen_monto", 0)
        c5.number_format  = FMT_MONEY
        c5.font           = Font(name="Arial", size=9, bold=True, color=monto_fg)
        c5.fill           = fill(row_bg)
        c5.alignment      = Alignment(horizontal="right", vertical="center")
        c5.border         = border_bottom()

        # Margen %
        pct_val = item.get("margen_pct", 0) / 100
        c6 = ws.cell(row=r, column=6)
        c6.value         = pct_val
        c6.number_format = FMT_PCT_1
        c6.font          = Font(name="Arial", size=9, bold=True,
                                color=CLR_ROJO_FG if negat else
                                (CLR_VERDE_FG if item.get("margen_pct",0) >= alert_yellow else
                                 CLR_AMARILLO_FG if item.get("margen_pct",0) >= alert_red else CLR_ROJO_FG))
        c6.fill          = fill(row_bg)
        c6.alignment     = Alignment(horizontal="center", vertical="center")
        c6.border        = border_bottom()

        data_cell(7, item.get("unidades_vendidas", 0), fmt=FMT_MONEY_0, align="right")

        # Fórmulas calculadas en Excel
        precio_col  = get_column_letter(3)
        costo_col   = get_column_letter(4)
        uds_col     = get_column_letter(7)
        ing_col     = get_column_letter(8)
        cos_col     = get_column_letter(9)

        ws.cell(r, 8).value          = f"={precio_col}{r}*{uds_col}{r}"
        ws.cell(r, 8).number_format  = FMT_MONEY
        ws.cell(r, 8).font           = body_font(color="0F172A")
        ws.cell(r, 8).fill           = fill(row_bg)
        ws.cell(r, 8).alignment      = Alignment(horizontal="right", vertical="center")
        ws.cell(r, 8).border         = border_bottom()

        ws.cell(r, 9).value          = f"={costo_col}{r}*{uds_col}{r}"
        ws.cell(r, 9).number_format  = FMT_MONEY
        ws.cell(r, 9).font           = body_font(color="0F172A")
        ws.cell(r, 9).fill           = fill(row_bg)
        ws.cell(r, 9).alignment      = Alignment(horizontal="right", vertical="center")
        ws.cell(r, 9).border         = border_bottom()

        ws.cell(r, 10).value         = f"={ing_col}{r}-{cos_col}{r}"
        ws.cell(r, 10).number_format = FMT_MONEY
        ws.cell(r, 10).font          = Font(name="Arial", size=9, bold=True,
                                            color=CLR_VERDE_FG)
        ws.cell(r, 10).fill          = fill(row_bg)
        ws.cell(r, 10).alignment     = Alignment(horizontal="right", vertical="center")
        ws.cell(r, 10).border        = border_bottom()

        # Semáforo texto
        sem_label = ("⬛ Negativo" if negat else
                     sem_styles.get(sem, ("", "", "—"))[2])
        data_cell(11, sem_label, align="center",
                  fg=(CLR_NEGATIVO_FG if negat else sem_styles.get(sem, ("","","","0F172A"))[1]))

        ws.row_dimensions[r].height = 18

    # ── Fila de totales ────────────────────────────────────────────────────
    total_row = 5 + len(margenes)
    data_row_start = 5
    data_row_end   = total_row - 1

    total_style = {"fill": fill(CLR_TOTAL_BG), "border": border_thin()}

    for col in range(1, n_cols + 1):
        c = ws.cell(row=total_row, column=col)
        c.fill   = fill(CLR_TOTAL_BG)
        c.border = border_thin()
        c.font   = Font(name="Arial", size=9, bold=True, color=CLR_HEADER_BG)
        c.alignment = Alignment(horizontal="right" if col > 2 else "left", vertical="center")

    ws.cell(total_row, 1).value = "TOTALES"
    ws.cell(total_row, 7).value = f"=SUM(G{data_row_start}:G{data_row_end})"
    ws.cell(total_row, 7).number_format = FMT_MONEY_0
    ws.cell(total_row, 8).value = f"=SUM(H{data_row_start}:H{data_row_end})"
    ws.cell(total_row, 8).number_format = FMT_MONEY
    ws.cell(total_row, 9).value = f"=SUM(I{data_row_start}:I{data_row_end})"
    ws.cell(total_row, 9).number_format = FMT_MONEY
    ws.cell(total_row, 10).value = f"=SUM(J{data_row_start}:J{data_row_end})"
    ws.cell(total_row, 10).number_format = FMT_MONEY
    ws.row_dimensions[total_row].height = 22

    # Margen % promedio ponderado en el total
    ws.cell(total_row, 6).value = (
        f"=IFERROR(J{total_row}/H{total_row},0)"
    )
    ws.cell(total_row, 6).number_format = FMT_PCT_1


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    raw = sys.stdin.read()
    data = json.loads(raw)

    margenes     = data.get("margenes", [])
    resumen      = data.get("resumen", {})
    from_        = data.get("from", "")
    to_          = data.get("to", "")
    cost_field   = data.get("cost_field", "COS_PRO_UN")
    client_name  = data.get("client_name", "Cliente")
    currency     = data.get("currency", "$")
    iva_rate     = float(data.get("iva_rate", 16))
    alert_red    = float(data.get("alert_red", 10))
    alert_yellow = float(data.get("alert_yellow", 20))
    output_path  = data.get("output_path", "/tmp/margenes.xlsx")
    excluir_iva  = resumen.get("iva_excluido", False)

    wb = Workbook()

    # Hoja 1: Resumen Bono
    ws1 = wb.active
    build_hoja_resumen(ws1, resumen, client_name, from_, to_,
                       cost_field, currency, iva_rate)

    # Hoja 2: Detalle Márgenes
    ws2 = wb.create_sheet()
    build_hoja_margenes(ws2, margenes, currency,
                        alert_red, alert_yellow, excluir_iva, iva_rate)

    wb.save(output_path)
    print(f"OK:{output_path}", flush=True)


if __name__ == "__main__":
    main()
