#!/usr/bin/env python3
"""
storage/scripts/generate_inventario.py
Reporte consolidado de inventario — 3 hojas.
Recibe JSON por stdin. Dependencia: pip install openpyxl
"""
import sys, json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
C = {
    'dark':    '0F172A', 'white': 'FFFFFF', 'muted': '94A3B8',
    'critico': 'FEE2E2', 'critico_fg': 'B91C1C',
    'bajo':    'FEF3C7', 'bajo_fg':    '92400E',
    'alerta':  'E0F2FE', 'alerta_fg':  '0C4A6E',
    'ok':      'DCFCE7', 'ok_fg':      '14532D',
    'parcial': 'FEF3C7', 'parcial_fg': '92400E',
    'border':  'E2E8F0', 'total_bg':   'F1F5F9',
    'header2': '1E293B',
}

def fill(h): return PatternFill('solid', fgColor=h)
def font(size=10, bold=False, color='0F172A'): return Font(name='Arial', size=size, bold=bold, color=color)
def border():
    s = Side(style='thin', color=C['border'])
    return Border(left=s, right=s, top=s, bottom=s)
def border_b():
    return Border(bottom=Side(style='thin', color=C['border']))
def align(h='left', wrap=False): return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

FMT_MONEY = '#,##0.00'
FMT_NUM   = '#,##0'
FMT_PCT   = '0.0%'

def title_row(ws, row, text, n_cols, bg=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row, 1)
    c.value = text
    c.font = Font(name='Arial', size=12, bold=True, color=C['white'])
    c.fill = fill(bg or C['dark'])
    c.alignment = align('center')
    ws.row_dimensions[row].height = 28

def header_row(ws, row, cols):
    for i, (label, width) in enumerate(cols, 1):
        c = ws.cell(row, i)
        c.value = label
        c.font = font(9, bold=True, color=C['white'])
        c.fill = fill(C['header2'])
        c.alignment = align('center')
        c.border = border()
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 20

# ── HOJA 1: Stock Crítico ─────────────────────────────────────────────────────
def hoja_stock(ws, data, client, from_, to_):
    ws.title = 'Stock Crítico'
    ws.sheet_view.showGridLines = False

    cols = [
        ('Código', 14), ('Descripción', 40), ('Stock Actual', 14),
        ('Stock Mínimo', 14), ('Comprometido', 13), ('Stock Libre', 12),
        ('% Cubierto', 11), ('Déficit', 12), ('Nivel', 11),
    ]
    title_row(ws, 1, f'{client} — Stock Crítico | Corte: {to_}', len(cols))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    ws.cell(2,1).value = f'Artículos con stock ≤ mínimo configurado. Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws.cell(2,1).font  = font(9, color=C['muted'])
    ws.cell(2,1).alignment = align('center')
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 8

    header_row(ws, 4, cols)
    ws.freeze_panes = 'A5'

    nivel_map = {
        'critico': (C['critico'], C['critico_fg'], '🔴 Crítico'),
        'bajo':    (C['bajo'],    C['bajo_fg'],    '🟡 Bajo'),
        'alerta':  (C['alerta'],  C['alerta_fg'],  '🔵 Alerta'),
    }

    for r_off, item in enumerate(data, 5):
        nivel = item.get('nivel', 'alerta')
        bg, fg, lbl = nivel_map.get(nivel, (C['alerta'], C['alerta_fg'], ''))
        pct = item.get('pct_cubierto', 0) / 100

        vals = [
            item.get('codigo',''),
            item.get('descripcion',''),
            item.get('stock_actual', 0),
            item.get('stock_minimo', 0),
            item.get('stock_comprometido', 0),
            item.get('stock_libre', 0),
            pct,
            item.get('deficit', 0),
            lbl,
        ]
        fmts = [None, None, FMT_NUM, FMT_NUM, FMT_NUM, FMT_NUM, FMT_PCT, FMT_NUM, None]
        aligns = ['left','left','right','right','right','right','center','right','center']

        for ci, (v, fmt, al) in enumerate(zip(vals, fmts, aligns), 1):
            c = ws.cell(r_off, ci)
            c.value = v
            c.font  = font(9, bold=(ci==1), color=fg)
            c.fill  = fill(bg)
            c.alignment = align(al)
            c.border = border_b()
            if fmt: c.number_format = fmt
        ws.row_dimensions[r_off].height = 17

    # Totales
    tr = 5 + len(data)
    ws.cell(tr, 1).value = f'TOTAL: {len(data)} artículos'
    ws.cell(tr, 1).font  = font(9, bold=True)
    ws.cell(tr, 1).fill  = fill(C['total_bg'])
    ws.cell(tr, 1).border= border()
    for ci in range(2, len(cols)+1):
        ws.cell(tr, ci).fill = fill(C['total_bg'])
        ws.cell(tr, ci).border = border()


# ── HOJA 2: Salidas No Comerciales ───────────────────────────────────────────
def hoja_salidas(ws, data, currency, from_, to_):
    ws.title = 'Salidas No Comerciales'
    ws.sheet_view.showGridLines = False

    cols = [
        ('N° Ajuste', 14), ('Fecha', 12), ('Código Art.', 14),
        ('Descripción', 38), ('Tipo Movimiento', 20),
        ('Cantidad', 11), ('Costo Estim.', 14), ('Riesgo', 12),
    ]
    title_row(ws, 1, f'Salidas No Comerciales | {from_} → {to_}', len(cols), bg='7F1D1D')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    ws.cell(2,1).value = f'Ajustes de inventario que no corresponden a ventas. Total: {len(data)} movimientos.'
    ws.cell(2,1).font  = font(9, color=C['muted'])
    ws.cell(2,1).alignment = align('center')
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 8

    header_row(ws, 4, cols)
    ws.freeze_panes = 'A5'

    riesgo_bg = {5:'FEE2E2', 4:'FEE2E2', 3:'FEF3C7', 2:'E0F2FE', 1:'DCFCE7'}
    riesgo_fg = {5:'7F1D1D', 4:'B91C1C', 3:'92400E', 2:'0C4A6E', 1:'14532D'}

    for r_off, item in enumerate(data, 5):
        r = item.get('riesgo', 2)
        row_bg = riesgo_bg.get(r, 'FFFFFF') if r >= 4 else 'FFFFFF'
        r_fg   = riesgo_fg.get(r, '0F172A')

        vals = [
            item.get('numero_ajuste',''), item.get('fecha',''),
            item.get('articulo_codigo',''), item.get('articulo_descripcion',''),
            item.get('tipo_label', item.get('tipo_movimiento','')),
            abs(item.get('cantidad', 0)),
            item.get('costo_estimado', 0),
            item.get('riesgo_label', ''),
        ]
        fmts  = [None,None,None,None,None,FMT_NUM,FMT_MONEY,None]
        aligns= ['left','center','left','left','left','right','right','center']
        bolds = [False,False,True,False,False,False,True,False]

        for ci, (v, fmt, al, bd) in enumerate(zip(vals,fmts,aligns,bolds), 1):
            c = ws.cell(r_off, ci)
            c.value = v
            c.font  = font(9, bold=bd, color=r_fg if (r >= 4 and ci not in [3,4]) else '0F172A')
            c.fill  = fill(row_bg)
            c.alignment = align(al)
            c.border = border_b()
            if fmt: c.number_format = fmt
        ws.row_dimensions[r_off].height = 17

    # Total costo
    tr = 5 + len(data)
    ws.cell(tr,1).value = 'COSTO TOTAL ESTIMADO'
    ws.cell(tr,1).font  = font(9, bold=True, color=C['critico_fg'])
    ws.cell(tr,1).fill  = fill('FEE2E2')
    ws.cell(tr,1).border= border()
    costo_total_col = get_column_letter(7)
    ws.cell(tr,7).value = f'=SUM({costo_total_col}5:{costo_total_col}{tr-1})'
    ws.cell(tr,7).number_format = FMT_MONEY
    ws.cell(tr,7).font  = font(9, bold=True, color=C['critico_fg'])
    ws.cell(tr,7).fill  = fill('FEE2E2')
    ws.cell(tr,7).alignment = align('right')
    ws.cell(tr,7).border = border()
    for ci in [2,3,4,5,6,8]:
        ws.cell(tr,ci).fill = fill('FEE2E2')
        ws.cell(tr,ci).border = border()
    ws.row_dimensions[tr].height = 20


# ── HOJA 3: Entradas vs Compras ───────────────────────────────────────────────
def hoja_entradas(ws, data, from_, to_):
    ws.title = 'Entradas vs Compras'
    ws.sheet_view.showGridLines = False

    cols = [
        ('N° Orden', 14), ('Fecha', 12), ('Proveedor', 28),
        ('Cód. Art.', 12), ('Descripción', 36),
        ('Cant. Ordenada', 14), ('Cant. Recibida', 14),
        ('Diferencia', 12), ('% Recibido', 12), ('Estado', 14),
    ]
    title_row(ws, 1, f'Entradas vs Órdenes de Compra | {from_} → {to_}', len(cols), bg='1E3A5F')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    ws.cell(2,1).value = f'Cruce de órdenes de compra contra mercancía efectivamente recibida. Total: {len(data)} líneas.'
    ws.cell(2,1).font  = font(9, color=C['muted'])
    ws.cell(2,1).alignment = align('center')
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 8

    header_row(ws, 4, cols)
    ws.freeze_panes = 'A5'

    alerta_style = {
        'ok':          (C['ok'],      C['ok_fg'],       '✅ Completo'),
        'sin_entrada': ('EDE9FE',     '5B21B6',          '⛔ Sin entrada'),
        'critico':     (C['critico'], C['critico_fg'],   '🔴 Crítico'),
        'parcial':     (C['parcial'], C['parcial_fg'],   '🟡 Parcial'),
        'leve':        (C['alerta'],  C['alerta_fg'],    '🔵 Leve'),
    }

    for r_off, item in enumerate(data, 5):
        alerta = item.get('alerta', 'leve')
        bg, fg, lbl = alerta_style.get(alerta, (C['alerta'], C['alerta_fg'], alerta))
        pct = item.get('pct_recibido', 0) / 100

        vals = [
            item.get('numero_orden',''), item.get('fecha',''),
            item.get('proveedor',''), item.get('articulo_codigo',''),
            item.get('articulo_descripcion',''),
            item.get('cantidad_ordenada',0), item.get('cantidad_recibida',0),
            item.get('diferencia',0), pct, lbl,
        ]
        fmts  = [None,None,None,None,None,FMT_NUM,FMT_NUM,FMT_NUM,FMT_PCT,None]
        aligns= ['left','center','left','left','left','right','right','right','center','center']
        bolds = [False,False,False,True,False,False,False,True,True,True]
        use_bg = alerta in ('critico','sin_entrada')

        for ci, (v,fmt,al,bd) in enumerate(zip(vals,fmts,aligns,bolds),1):
            c = ws.cell(r_off, ci)
            c.value = v
            c.font  = font(9, bold=bd,
                           color=(fg if (use_bg or (bd and ci in [8,9,10])) else '0F172A'))
            c.fill  = fill(bg if use_bg else 'FFFFFF')
            c.alignment = align(al)
            c.border = border_b()
            if fmt: c.number_format = fmt
        # Colorear siempre estado y pct
        ws.cell(r_off, 9).fill  = fill(bg)
        ws.cell(r_off, 10).fill = fill(bg)
        ws.row_dimensions[r_off].height = 17

    # Totales
    tr = 5 + len(data)
    ws.cell(tr,1).value = 'TOTALES'
    ws.cell(tr,1).font  = font(9, bold=True)
    ws.cell(tr,1).fill  = fill(C['total_bg'])
    ws.cell(tr,1).border= border()
    for ci, col_idx in [(6,6),(7,7),(8,8)]:
        cl = get_column_letter(col_idx)
        ws.cell(tr,ci).value = f'=SUM({cl}5:{cl}{tr-1})'
        ws.cell(tr,ci).number_format = FMT_NUM
        ws.cell(tr,ci).font = font(9, bold=True)
        ws.cell(tr,ci).fill = fill(C['total_bg'])
        ws.cell(tr,ci).alignment = align('right')
        ws.cell(tr,ci).border = border()
    for ci in [2,3,4,5,9,10]:
        ws.cell(tr,ci).fill = fill(C['total_bg'])
        ws.cell(tr,ci).border = border()
    # % global
    ws.cell(tr,9).value = f'=IFERROR(G{tr}/F{tr},0)'
    ws.cell(tr,9).number_format = FMT_PCT
    ws.cell(tr,9).font = font(9, bold=True)
    ws.cell(tr,9).fill = fill(C['total_bg'])
    ws.cell(tr,9).alignment = align('center')
    ws.cell(tr,9).border = border()
    ws.row_dimensions[tr].height = 20


# ── ENTRY POINT ───────────────────────────────────────────────────────────────
def main():
    data        = json.loads(sys.stdin.read())
    stock       = data.get('stock', [])
    salidas     = data.get('salidas', [])
    entradas    = data.get('entradas', [])
    from_       = data.get('from','')
    to_         = data.get('to','')
    client      = data.get('client_name','Cliente')
    currency    = data.get('currency','$')
    output_path = data.get('output_path','/tmp/inventario.xlsx')

    wb = Workbook()

    ws1 = wb.active
    hoja_stock(ws1, stock, client, from_, to_)

    ws2 = wb.create_sheet()
    hoja_salidas(ws2, salidas, currency, from_, to_)

    ws3 = wb.create_sheet()
    hoja_entradas(ws3, entradas, from_, to_)

    wb.save(output_path)
    print(f'OK:{output_path}', flush=True)

if __name__ == '__main__':
    main()
